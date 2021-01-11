from .models import Shipper, Line, Vessel, Consignment, ExFile
from django.contrib.auth.models import User

from .serializers import (ShipperSerializer,
                          LineSerializer, VesselSerializer, ConsignmentSerializer, UserSerializer, ExFileSerializer)
from .filters import (ShipperFilter,
                      LineFilter, VesselFilter, ConsignmentFilter, UserFilter)

from rest_framework import status, viewsets
from rest_framework.response import Response
from rest_framework.parsers import FileUploadParser

# Excel...
import openpyxl

from datetime import datetime

import os
from django.conf import settings


class UserViewSet(viewsets.ModelViewSet):
    queryset = User.objects.all()
    serializer_class = UserSerializer
    filter_class = UserFilter


class ShipperViewSet(viewsets.ModelViewSet):
    queryset = Shipper.objects.all()
    serializer_class = ShipperSerializer
    filter_class = ShipperFilter


class LineViewSet(viewsets.ModelViewSet):
    queryset = Line.objects.all()
    serializer_class = LineSerializer
    filter_class = LineFilter


class VesselViewSet(viewsets.ModelViewSet):
    queryset = Vessel.objects.all()
    serializer_class = VesselSerializer
    filter_class = VesselFilter


class ConsignmentViewSet(viewsets.ModelViewSet):
    queryset = Consignment.objects.all()
    serializer_class = ConsignmentSerializer
    filter_class = ConsignmentFilter


class ExFileViewSet(viewsets.ModelViewSet):
    queryset = ExFile.objects.all()
    serializer_class = ExFileSerializer
    parser_class = (FileUploadParser,)

    def create(self, request):
        serializer = ExFileSerializer(data=request.data)
        if serializer.is_valid():
            serializer.save()

            # TODO: Format DEPARTURE and ARRIVAL dates properly in excel. YYYY-MM-DD
            # TODO: Probably working well for in spite of inappropriate format.
            try:
                new_path = os.path.abspath(os.path.join(settings.MEDIA_ROOT, '..', ) + serializer.data.get('file'))
                wb = openpyxl.load_workbook(new_path, data_only=True)  # Option To read the data only...no formulae
                sheet = wb.active

                # Heading
                # CONTAINER  SHIPPER  LINE  DEPARTURE  ARRIVAL  VESSEL  COMPLETE  RECEIPT NUMBER
                head_dict = {}
                head_key_list, head_val_list = [], []
                for r in range(1, 2, 1):
                    for c in range(1, sheet.max_column + 1):
                        head_dict[c] = str(sheet.cell(row=r, column=c).value).strip().lower()

                head_val_list = list(head_dict.values())
                print('-----all headers....-----', head_dict)

                # Body
                # TTNU8492121  JOSEPHINE HAPAG  2020-03-08 00:00:00  2020-04-13 00:00:00  GENOA EXPRESS  YES
                proj_dict = {}
                shipper_first_name, line_name = '', ''
                _shipper, _line, _vessel, departure, arrival = None, None, None, None, None
                complete, receipt_number = None, None
                _departure, _arrival = None, None

                for row in range(2, sheet.max_row):
                    for col in range(1, sheet.max_column + 1):

                        # Add Shipper
                        # TODO: Not enough information in excel to add shipper properly
                        if col == (head_val_list.index('shipper') + 1):
                            shipper_first_name = sheet.cell(row=row, column=col).value
                            shipper_first_name = shipper_first_name.strip() if shipper_first_name is not None else ''
                            _shipper, created = Shipper.objects.get_or_create(first_name=shipper_first_name,
                                                                              modified_by=request.user)

                        # Add Line
                        if col == (head_val_list.index('line') + 1):
                            line_name = sheet.cell(row=row, column=col).value
                            line_name = line_name.strip() if line_name is not None else ''
                            _line, created = Line.objects.get_or_create(name=line_name, modified_by=request.user)

                        # Add Vessel
                        if col == (head_val_list.index('vessel') + 1):
                            vessel_name = sheet.cell(row=row, column=col).value
                            vessel_name = vessel_name.strip() if vessel_name is not None else ''
                            _vessel, created = Vessel.objects.get_or_create(name=vessel_name, line=_line,
                                                                            modified_by=request.user)
                        # Departure
                        if col == (head_val_list.index('departure') + 1):
                            departure = sheet.cell(row=row, column=col).value
                            # departure = departure.strip() if departure is not None else ''

                        # Arrival
                        if col == (head_val_list.index('arrival') + 1):
                            arrival = sheet.cell(row=row, column=col).value
                            # arrival = arrival.strip() if arrival is not None else ''

                        # Complete
                        if col == (head_val_list.index('complete') + 1):
                            complete = sheet.cell(row=row, column=col).value
                            complete = complete.strip() if complete is not None else ''

                        # Receipt Number
                        if col == (head_val_list.index('reciept number') + 1):
                            receipt_number = sheet.cell(row=row, column=col).value
                            receipt_number = receipt_number.strip() if receipt_number is not None else ''

                        # Add Consignment
                        if col == (head_val_list.index('container') + 1):
                            container = sheet.cell(row=row, column=col).value
                            container = container.strip() if container is not None else ''

                        if container and col == sheet.max_column:
                            _departure = departure.date() if isinstance(departure, datetime) else None
                            _arrival = arrival.date() if isinstance(arrival, datetime) else None

                            if not Consignment.objects.filter(container=container).exists():
                                _consignment = Consignment(container=container,
                                                           shipper=_shipper,
                                                           line=_line,
                                                           departure=_departure,
                                                           arrival=_arrival,
                                                           vessel=_vessel,
                                                           status=complete,
                                                           receipt_number=receipt_number,
                                                           modified_by=request.user)
                                _consignment.save()

            except Exception as ex:
                print("------- Exception ---------")
                print(ex)

            return Response(serializer.data)